import os, xlrd, configparser

import openpyxl as op
from decimal import *
from skuproject.settings import STATIC_ROOT, STATIC_URL
from .models import UploadedBaseInfo, SKU, Capsule, Season, Size


def get_size_short(sizelong):
    '''
    sizelong: size like "140*72*63"
    return sizeshort: short version of size, like "140"
    if it's DOUBLE SIZE (like "140*146") - returns sizeshort "140*146"
    '''

    if sizelong.count('*') == 2:    # if there are two "*"
        t = sizelong.find('*')
        sizeshort = sizelong[0:t]
        case = 'NORMAL SIZE'

    elif sizelong.count('*') == 1:  # if there are one "*"
        t = sizelong.find('*')      # there are two possible options: DOUBLE SIZE or TODDLER SIZE
        temp_size_divided = sizelong.split('*')
        try:
            left_part_of_size = int(temp_size_divided[0])
            right_part_of_size = int(temp_size_divided[1])

            if left_part_of_size > right_part_of_size:
                sizeshort = sizelong[0:t]   # if size like "92*52" we'll leave just "92"
                case = 'TODDLER SIZE'
            else:
                sizeshort = sizelong    # if "14*16" - leave no change
                case = 'DOUBLE SIZE'
        except:
            case = "OTHER"  # other variants like "6*6S" etc.
            sizeshort = sizelong

    elif sizelong == 'No size':
        sizeshort = 'No size'
        case = 'NO SIZE'

    elif sizelong.isdigit():
        case = 'ONE DIGIT SIZE'
        sizeshort = sizelong

    else:
        case = "OTHER"  # All other variants like "120/30" etc.
        sizeshort = sizelong

    return sizeshort


def string_to_decimal(string):
    try:
        string = string.replace(' ', '')
        return Decimal(string)
    except:
        return 0


def string_to_integer(string):
    try:
        string = string.replace(' ', '')
        pos = string.find('.')
        if pos != -1:
            return int(string[:pos])
        return int(string)
    except:
        return 0


def sheet_cell(row, column):
    '''
    One universal function for get sheet.cell value, both for .xlsx, and for .xls
    For .xlsx - I use openpyxl, for .xls - xlrd
    '''
    global sheet, file_type
    if file_type == 'XLS':
        return sheet.cell_value(row, column)
    elif file_type == 'XLSX':
        return sheet.cell(row=row, column=column).value


def handle_uploaded_file(excel_file, user_id):
    '''
    This function is parsing uploaded file and write it to database.
    Version with support of both xls and xlsx types of file
    Also with support of NEW versions of 1C where is no Номенклатура.Код column
    Номенклатура looks like "22001GMC1205-98*52*48 Футболка", so sku, size and name are in one cell
    '''
    global sheet, file_type
    file_type = ''
    filename = excel_file.name.lower()
    if filename.endswith('.xlsx'):
        file_type = 'XLSX'
    elif filename.endswith('.xls'):
        file_type = 'XLS'
    else:
        file_type = 'UNKNOWN'

    # Reading config file
    seasons_configlist = []
    capsules_configlist = []
    error_message = ''
    seasonname_from_config = ''
    capsulename_from_config = ''
    try:
        config = configparser.ConfigParser()
        config.read('sku_config.ini')
        seasons_configlist = config['Seasons']
        capsules_configlist = config['Capsules']
    except:
        print("Файл конфигурации 'sku_config.ini' не обнаружен. Названия коллекций могут не подгружаться.")
        return False, "Файл конфигурации 'sku_config.ini' не обнаружен."

    # Reading Excel file
    try:
        if file_type == 'XLS':
            rb = xlrd.open_workbook(file_contents=excel_file.read())
            sheet = rb.sheet_by_index(0)
            rows_quantity = sheet.nrows
            cols_quantity = sheet.ncols

            range_start_rows = 0  # In xlrd indexing starts from 0
            range_end_rows = rows_quantity
            range_start_cols = 0
            range_end_cols = cols_quantity

        elif file_type == 'XLSX':
            wb = op.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            rows_quantity = sheet.max_row
            cols_quantity = sheet.max_column

            range_start_rows = 1  # In openpyxl indexing starts from 1
            range_end_rows = rows_quantity + 1
            range_start_cols = 1
            range_end_cols = cols_quantity + 1
        else:
            return False, 'Неизвестный формат файла'
    except:
        return False, 'Если вы уверены что отчет из 1С выгружен правильно, попробуйте открыть его в Экселе, изменить в нем что-нибудь, (ширину столбца, цвет ячейки и т.д.) и сохранить. Затем загрузите его сюда. Со старыми версиями 1С бывают такие ситуации'

    # Parsing the header
    row_header = -1
    period_from_xls = ''
    col_sku = col_name = col_quantity_sold = col_sellsumm_sold = -1
    col_sellprice_sold = col_costsumm_sold = col_costprice_sold = -1
    col_income = col_quantity_instock = col_costsumm_instock = -1
    row_sku = row_name = -1
    shift = 0

    for rownum in range(range_start_rows, range_end_rows):
        for colnum in range(range_start_cols, range_end_cols):

            if str(sheet_cell(rownum, colnum))[:7] == 'Период:':
                t1 = sheet_cell(rownum, colnum).find('\n')  # find line break
                period_from_xls = sheet_cell(rownum, colnum)[8:t1]

            if sheet_cell(rownum, colnum) == 'Номенклатура.Код':
                col_sku = colnum
                row_sku = rownum

            if sheet_cell(rownum, colnum) == 'Номенклатура':
                col_name = colnum
                row_name = rownum

            if sheet_cell(rownum, colnum) == 'Кол-во (продажи)':
                col_quantity_sold = colnum
                row_header = rownum

            if sheet_cell(rownum, colnum) == 'Сумма, руб. (продажи)':
                col_sellsumm_sold = colnum

            # if sheet.cell_value(rownum, colnum) == 'Цена розничная, руб. (продажи)':
            #     col_sellprice_sold=colnum

            if sheet_cell(rownum, colnum) == 'Сумма себестоимость, руб. (продажи)':
                col_costsumm_sold = colnum

            # if sheet.cell_value(rownum, colnum) == 'Цена себестоимость, руб. (продажи)':
            #     col_costprice_sold=colnum

            if sheet_cell(rownum, colnum) == 'Доход, руб.':
                col_income = colnum

            if sheet_cell(rownum, colnum) == 'Кол-во (остатки)':
                col_quantity_instock = colnum

            if sheet_cell(rownum, colnum) == 'Сумма себестоимость, руб. (остатки)':
                col_costsumm_instock = colnum

        if row_header != -1 and rownum > row_header+2:
            break  # No need to go through whole file till the end

    if col_name == -1:
        return False, "Нет столбца 'Номенклатура'"
    if col_quantity_sold == -1:
        return False, "Нет столбца 'Кол-во (продажи)'"
    if col_sellsumm_sold == -1:
        return False, "Нет столбца 'Сумма, руб. (продажи)'"
    if col_costsumm_sold == -1:
        return False, "Нет столбца 'Сумма себестоимость, руб. (продажи)'"
    if col_income == -1:
        return False, "Нет столбца 'Доход, руб.'"
    if col_quantity_instock == -1:
        return False, "Нет столбца 'Кол-во (остатки)'"
    if col_costsumm_instock == -1:
        return False, "Нет столбца 'Сумма себестоимость, руб. (остатки)'"

    if col_sku == -1:  # There is no "Номенклатура.Код" column
        print('Нет столбца Номенклатура.Код')
        # Two possible cases:
        # 1) version_1c='OLD' and just forgot to include Номенклатура.Код column. Alert an error.
        # 2) version_1c='NEW', so sku and name are in Номенклатура column
        it_has_sku = False

        row = row_header + 1  # Starting just below header

        while row < range_end_rows:
            work_cell = sheet_cell(row, col_name)  # Check Номенклатура column
            if work_cell:
                work_cell.strip()

            # Simple validation: if first 3 symbols are digits - this is sku
            if str(work_cell)[:3].isdigit():
                it_has_sku = True
                break  #
            else:
                row += 1
                continue

        if it_has_sku:
            version_1c = 'NEW'
            col_sku = col_name
        else:
            return False, 'Нет столбца Номенклатура.Код и в столбце Номенклатура не найдены артикулы'

    else:
        version_1c = 'OLD'

    print('Version 1C: ', version_1c)

    if version_1c == 'OLD':
        shift = row_name - row_sku      # Номенклатура and Номенклатура.Код can be on different rows
                                        # shift - it's relative row position between them
    print ('shift: ', shift)

    # Forming lists first, then bulk_create
    season_list = []  # lists of instances
    capsule_list = []
    sku_list = []
    size_list = []

    seasons_checklist = []  # lists for checking if the instance already exist
    capsules_checklist = []
    skus_checklist = []

    row = row_header + 1  # Starting just below header

    while row < range_end_rows:
        work_cell = sheet_cell(row, col_sku)
        if work_cell:
            work_cell.strip()

        # Simple validation: if first 3 symbols are digits - this is sku
        if not str(work_cell)[:3].isdigit():
            row += 1
            continue

        # Make correct values types
        cell = sheet_cell(row, col_quantity_sold)  # Количество (продажи)
        if cell:
            cell = str(cell)
            q_s = string_to_integer(cell)
        else:
            q_s = 0

        cell = sheet_cell(row, col_sellsumm_sold)  # Сумма (продажи)
        if cell:
            cell = str(cell)
            ss_s = string_to_decimal(cell)
        else:
            ss_s = 0

        cell = sheet_cell(row, col_costsumm_sold)  # Сумма себестоимость (продажи)
        if cell:
            cell = str(cell)
            cs_s = string_to_decimal(cell)
        else:
            cs_s = 0

        cell = sheet_cell(row, col_income)  # Доход
        if cell:
            cell = str(cell)
            incm = string_to_decimal(cell)
        else:
            incm = 0

        cell = sheet_cell(row, col_quantity_instock)  # Количество (остатки)
        if cell:
            cell = str(cell)
            q_i = string_to_integer(cell)
        else:
            q_i = 0

        cell = sheet_cell(row, col_costsumm_instock)  # Сумма себестоимость (остатки)
        if cell:
            cell = str(cell)
            cs_i = string_to_decimal(cell)
        else:
            cs_i = 0

        # Checking values:
        if q_s <= 0 or ss_s <= 0 or cs_s <= 0:  # if quantity or summ <=0, everything else doesn't make sense
            q_s = 0
            ss_s = 0
            cs_s = 0
            incm = 0

            if q_i <= 0 or cs_i <= 0:  # if both SOLD and INSTOCK  <=0, skip row
                row += 1
                continue

        if q_i <= 0 or cs_i <= 0:
            q_i = cs_i = 0

        # Parsing sku
        # Recieve 1) sku_nosize, 2) sizelong, 3) sizeshort 4) sku_name (empty is possible) 5) season 6) capsule
        sku_name = ''
        t = work_cell.find('-')

        if t != -1:
            sku_nosize = work_cell[:t]
            sizelong = work_cell[t + 1:]
        else:
            sku_nosize = work_cell[:12]
            sizelong = 'No size'

        if version_1c == 'NEW' and t != -1:
            postn = sizelong.find(' ')

            if postn == -1:
                pass
                # sizelong=sizelong
                # sku_name=''

            elif sizelong[:postn].lower() == 'no':  # 'no size' there
                tmp = sizelong[postn + 1:]
                postn = tmp.find(' ')  # second space

                if postn != -1:
                    sizelong = 'No size'
                    sku_name = tmp[postn + 1:]

            else:
                sku_name = sizelong[postn + 1:]
                sizelong = sizelong[:postn]

        elif version_1c == 'OLD':
            sku_name = sheet_cell(row + shift, col_name)
        # print('sku_nosize', sku_nosize)

        season = work_cell[:5]
        if season[3:5] == 'GS' or season[3:5] == 'gs':
            pass  # SCHOOL, we take first 5 symbols
        else:
            season = season[0:3]  # NOT SCHOOL, we take first 3 symbols

        capsule = sku_nosize[:6]

        # Add Season
        if season not in seasons_checklist:
            seasons_checklist.append(season)
            try:
                seasonname_from_config = seasons_configlist[season]
            except:
                seasonname_from_config = '----'

            filename = 'images/' + season + '/' + season + '.jpg'
            image_season = STATIC_URL + filename

            if not os.path.isfile(STATIC_ROOT + '/' + filename):
                image_season = ''

            # I have to assign id by myself, because whole database firstly create in memory, and after all - bulk_create.
            # And I need to organize ForeignKey relations, so I need id for that.
            # So, my id is CharField and looks like "rownumber-user_id", i.e. 211-18.
            # And it's unique.
            string_id = str(row) + '-' + str(user_id.id)
            # print('string_id:', string_id)

            season_list.append(Season(season_firstletters=season,
                                      name=seasonname_from_config,
                                      img=image_season,
                                      user=user_id,
                                      id=string_id))

        # Add Capsule
        if capsule not in capsules_checklist:
            capsules_checklist.append(capsule)
            try:
                capsulename_from_config = capsules_configlist[capsule]
            except:
                capsulename_from_config = '----'

            filename = 'images/' + season + '/' + capsule + '.jpg'
            image_capsule = STATIC_URL + filename

            if not os.path.isfile(STATIC_ROOT + '/' + filename):
                image_capsule = ''

            season_id = [i.id for i in season_list if i.season_firstletters == season][0]

            string_id = str(row) + '-' + str(user_id.id)

            capsule_list.append(Capsule(capsule_firstletters=capsule,
                                        id=string_id,
                                        name=capsulename_from_config,
                                        img=image_capsule,
                                        user=user_id,
                                        season=Season(id=season_id)))

        # Add Sku
        if sku_nosize not in skus_checklist:
            skus_checklist.append(sku_nosize)

            filename = 'images/' + season + '/' + capsule + '/' + sku_nosize + '.jpg'
            image_sku = STATIC_URL + filename

            if not os.path.isfile(STATIC_ROOT + '/' + filename):
                image_sku = ''

            capsule_id = [i.id for i in capsule_list if i.capsule_firstletters == capsule][0]

            string_id = str(row) + '-' + str(user_id.id)

            sku_list.append(SKU(name=sku_name,
                                sku_firstletters=sku_nosize,
                                id=string_id,
                                capsule=Capsule(id=capsule_id),
                                user=user_id,
                                img=image_sku
                                ))

        # Add Size
        sku_id = [i.id for i in sku_list if i.sku_firstletters == sku_nosize][0]

        s_s = get_size_short(sizelong)

        size_list.append(Size(size_long=sizelong,
                              sku_full=work_cell,
                              sku=SKU(id=sku_id),
                              user=user_id,
                              quantity_sold=q_s,
                              sellsumm_sold=ss_s,
                              costsumm_sold=cs_s,
                              income=incm,
                              quantity_instock=q_i,
                              costsumm_instock=cs_i,
                              size_short=s_s
                              ))

        row += 1

    # #------------------Lets give rating badges---------------------
    # for _season in season_list:
    #     for _capsule in capsule_list:

    # -----------------Now lets write to database ----------------

    # Delete old base first
    UploadedBaseInfo.objects.filter(user=user_id).delete()
    Season.objects.filter(user=user_id).delete()
    Capsule.objects.filter(user=user_id).delete()
    SKU.objects.filter(user=user_id).delete()
    Size.objects.filter(user=user_id).delete()

    UploadedBaseInfo.objects.create(user=user_id, period=period_from_xls)
    Season.objects.bulk_create(season_list)
    Capsule.objects.bulk_create(capsule_list)
    SKU.objects.bulk_create(sku_list)
    Size.objects.bulk_create(size_list)

    return True, error_message


def handle_uploaded_file_OLD(excel_file, user_id):
    '''
    This function is parsing uploaded file and write it to database.
    Version supports only xls files.
    And without support of NEW versions of 1C (where column Номенклатура.Код isn't exists).
    '''

    # Reading config file
    seasons_configlist = []
    capsules_configlist = []
    error_message = ''
    seasonname_from_config = ''
    capsulename_from_config = ''
    try:
        config = configparser.ConfigParser()
        config.read('sku_config.ini')
        seasons_configlist = config['Seasons']
        capsules_configlist = config['Capsules']
    except:
        print("Файл конфигурации 'sku_config.ini' не обнаружен. Названия коллекций могут не подгружаться.")
        return False, "Файл конфигурации 'sku_config.ini' не обнаружен."

    # Reading Excel file
    rb = xlrd.open_workbook(file_contents=excel_file.read())
    sheet = rb.sheet_by_index(0)

    # Parsing header
    row_header = -1
    period_from_xls = ''
    col_sku = col_name = col_quantity_sold = col_sellsumm_sold = col_sellprice_sold = -1
    col_costsumm_sold = col_costprice_sold = -1
    col_income = col_quantity_instock = col_costsumm_instock = -1

    for rownum in range(sheet.nrows):
        for colnum in range(sheet.ncols):

            if str(sheet.cell_value(rownum, colnum))[:7] == 'Период:':
                t1 = sheet.cell_value(rownum, colnum).find('\n')  # find line break
                period_from_xls = sheet.cell_value(rownum, colnum)[8:t1]

            if sheet.cell_value(rownum, colnum) == 'Номенклатура.Код':
                col_sku = colnum

            if sheet.cell_value(rownum, colnum) == 'Номенклатура':
                col_name = colnum

            if sheet.cell_value(rownum, colnum) == 'Кол-во (продажи)':
                col_quantity_sold = colnum
                row_header = rownum

            if sheet.cell_value(rownum, colnum) == 'Сумма, руб. (продажи)':
                col_sellsumm_sold = colnum

            if sheet.cell_value(rownum, colnum) == 'Сумма себестоимость, руб. (продажи)':
                col_costsumm_sold = colnum

            if sheet.cell_value(rownum, colnum) == 'Доход, руб.':
                col_income = colnum

            if sheet.cell_value(rownum, colnum) == 'Кол-во (остатки)':
                col_quantity_instock = colnum

            if sheet.cell_value(rownum, colnum) == 'Сумма себестоимость, руб. (остатки)':
                col_costsumm_instock = colnum

        if row_header != -1 and rownum > row_header:
            break  # No need to go through whole file till the end

    if col_sku == -1:
        return False, "Нет столбца 'Номенклатура.Код'"
    if col_name == -1:
        return False, "Нет столбца 'Номенклатура'"
    if col_quantity_sold == -1:
        return False, "Нет столбца 'Кол-во (продажи)'"
    if col_sellsumm_sold == -1:
        return False, "Нет столбца 'Сумма, руб. (продажи)'"
    if col_costsumm_sold == -1:
        return False, "Нет столбца 'Сумма себестоимость, руб. (продажи)'"
    if col_income == -1:
        return False, "Нет столбца 'Доход, руб.'"
    if col_quantity_instock == -1:
        return False, "Нет столбца 'Кол-во (остатки)'"
    if col_costsumm_instock == -1:
        return False, "Нет столбца 'Сумма себестоимость, руб. (остатки)'"

    # Forming lists, later - bulk_create
    season_list = []  # lists of instances
    capsule_list = []
    sku_list = []
    size_list = []

    seasons_checklist = []  # # lists for checking if the instance already exist
    capsules_checklist = []
    skus_checklist = []

    row = row_header + 1  # start just below header

    while row < sheet.nrows:
        work_cell = sheet.cell_value(row, col_sku).strip()

        # Simple validation: if first 3 symbols are digits - this is sku
        if not str(work_cell)[:3].isdigit():
            row += 1
            continue

        # Make correct value types
        cell = sheet.cell_value(row, col_quantity_sold)  # Количество (продажи)
        if cell:
            cell = str(cell)
            q_s = string_to_integer(cell)
        else:
            q_s = 0

        cell = sheet.cell_value(row, col_sellsumm_sold)  # Сумма (продажи)
        if cell:
            cell = str(cell)
            ss_s = string_to_decimal(cell)
        else:
            ss_s = 0

        cell = sheet.cell_value(row, col_costsumm_sold)  # Сумма себестоимость (продажи)
        if cell:
            cell = str(cell)
            cs_s = string_to_decimal(cell)
        else:
            cs_s = 0

        cell = sheet.cell_value(row, col_income)  # Доход
        if cell:
            cell = str(cell)
            incm = string_to_decimal(cell)
        else:
            incm = 0

        cell = sheet.cell_value(row, col_quantity_instock)  # Количество (остатки)
        if cell:
            cell = str(cell)
            q_i = string_to_integer(cell)
        else:
            q_i = 0

        cell = sheet.cell_value(row, col_costsumm_instock)  # Сумма себестоимость (остатки)
        if cell:
            cell = str(cell)
            cs_i = string_to_decimal(cell)
        else:
            cs_i = 0

        # Checking values
        if q_s <= 0 or ss_s <= 0 or cs_s <= 0:  # if quantity or summ <=0, everything else doesn't make sense
            q_s = 0
            ss_s = 0
            cs_s = 0
            incm = 0

            if q_i <= 0 or cs_i <= 0:  # if both SOLD and INSTOCK  <=0, skip row
                row += 1
                continue

        if q_i <= 0 or cs_i <= 0:
            q_i = cs_i = 0

        # Parsing sku
        t = work_cell.find('-')
        if t != -1:
            sku_nosize = work_cell[:t]
            sizelong = work_cell[t + 1:]
        else:
            sku_nosize = work_cell
            sizelong = 'No size'

        season = work_cell[:5]
        if season[3:5] == 'GS' or season[3:5] == 'gs':
            pass  # SCHOOL, we take first 5 symbols
        else:
            season = season[0:3]  # NOT SCHOOL, we take first 3 symbols

        capsule = sku_nosize[:6]

        sku_name = sheet.cell_value(row + 1, col_name)

        # Add Season
        if season not in seasons_checklist:
            seasons_checklist.append(season)
            try:
                seasonname_from_config = seasons_configlist[season]
            except:
                seasonname_from_config = '----'

            filename = 'images/' + season + '/' + season + '.jpg'
            image_season = STATIC_URL + filename

            if not os.path.isfile(STATIC_ROOT + '/' + filename):
                image_season = ''

            # I have to assign id by myself, because whole database firstly create in memory, and after all - bulk_create.
            # And I need to organize ForeignKey relations, so I need id for that.
            # So, my id is CharField and looks like "rownumber-user_id", i.e. 211-18.
            # And it's unique.

            string_id = str(row) + '-' + str(user_id.id)
            print('string_id:', string_id)

            season_list.append(Season(season_firstletters=season,
                                      name=seasonname_from_config,
                                      img=image_season,
                                      user=user_id,
                                      id=string_id))

        # Add Capsule
        if capsule not in capsules_checklist:
            capsules_checklist.append(capsule)
            try:
                capsulename_from_config = capsules_configlist[capsule]
            except:
                capsulename_from_config = '----'

            filename = 'images/' + season + '/' + capsule + '.jpg'
            image_capsule = STATIC_URL + filename

            if not os.path.isfile(STATIC_ROOT + '/' + filename):
                image_capsule = ''

            season_id = [i.id for i in season_list if i.season_firstletters == season][0]

            string_id = str(row) + '-' + str(user_id.id)

            capsule_list.append(Capsule(capsule_firstletters=capsule,
                                        id=string_id,
                                        name=capsulename_from_config,
                                        img=image_capsule,
                                        user=user_id,
                                        season=Season(id=season_id)))

        # Add Sku
        if sku_nosize not in skus_checklist:
            skus_checklist.append(sku_nosize)

            filename = 'images/' + season + '/' + capsule + '/' + sku_nosize + '.jpg'
            image_sku = STATIC_URL + filename

            if not os.path.isfile(STATIC_ROOT + '/' + filename):
                image_sku = ''

            capsule_id = [i.id for i in capsule_list if i.capsule_firstletters == capsule][0]

            string_id = str(row) + '-' + str(user_id.id)

            sku_list.append(SKU(name=sku_name,
                                sku_firstletters=sku_nosize,
                                id=string_id,
                                capsule=Capsule(id=capsule_id),
                                user=user_id,
                                img=image_sku
                                ))

        # Add Size
        sku_id = [i.id for i in sku_list if i.sku_firstletters == sku_nosize][0]

        s_s = get_size_short(sizelong)

        size_list.append(Size(size_long=sizelong,
                              sku_full=work_cell,
                              sku=SKU(id=sku_id),
                              user=user_id,
                              quantity_sold=q_s,
                              sellsumm_sold=ss_s,
                              costsumm_sold=cs_s,
                              income=incm,
                              quantity_instock=q_i,
                              costsumm_instock=cs_i,
                              size_short=s_s
                              ))

        row += 1

    #   ------------------ Now lets write to database ----------------

    # Delete database first (only for current user of course)
    UploadedBaseInfo.objects.filter(user=user_id).delete()
    Season.objects.filter(user=user_id).delete()
    Capsule.objects.filter(user=user_id).delete()
    SKU.objects.filter(user=user_id).delete()
    Size.objects.filter(user=user_id).delete()

    # Then bulk_create
    UploadedBaseInfo.objects.create(user=user_id, period=period_from_xls)
    Season.objects.bulk_create(season_list)
    Capsule.objects.bulk_create(capsule_list)
    SKU.objects.bulk_create(sku_list)
    Size.objects.bulk_create(size_list)

    return True, error_message


def upload_onway_bill(xlsx_file, user_id):
    '''
    This function parses uploaded bill, then adds to database sizes, that are ON WAY, they aren't in 1C yet.
    (It's adding, without deleting existing database).
    Added sizes not taking part in calculations, because it's temporary. Sizes just figured in SKU page, user can see them.
    '''

    # Reading config file
    seasons_configlist = []
    capsules_configlist = []
    error_message = ''
    seasonname_from_config = ''
    capsulename_from_config = ''
    try:
        config = configparser.ConfigParser()
        config.read('sku_config.ini')
        seasons_configlist = config['Seasons']
        capsules_configlist = config['Capsules']
    except:
        print("Файл конфигурации 'sku_config.ini' не обнаружен. Названия коллекций могут не подгружаться.")
        return False, "Файл конфигурации 'sku_config.ini' не обнаружен."

    # Reading Excel file
    try:
        wb = op.load_workbook(xlsx_file, data_only=True)
        sheet = wb.active
        print('Файл открыт')
        print('Количество строк', sheet.max_row )
        print('Количество столбцов', sheet.max_column )

    except:
        return False, 'Не получается прочитать файл'

    # Parsing the header
    row_header = -1

    col_sku = col_name = col_quantity = col_summ = -1
    skuname = ''

    for rownum in range(1, sheet.max_row + 1):
        for colnum in range(1, sheet.max_column + 1):

            if str(sheet.cell(row=rownum, column=colnum).value) == 'Артикул':
                col_sku = colnum
                row_header = rownum

            if str(sheet.cell(row=rownum, column=colnum).value) == 'Товары (работы, услуги)':
                col_name = colnum

            if str(sheet.cell(row=rownum, column=colnum).value) == 'Кол-во':
                col_quantity = colnum

            if str(sheet.cell(row=rownum, column=colnum).value) == 'Сумма':
                col_summ = colnum

        if row_header != -1 and rownum > row_header:
            break  # No need to go through the whole file till the end

    if col_sku == -1:
        return False, "Нет столбца 'Артикул'"
    if col_name == -1:
        return False, "Нет столбца 'Товары (работы, услуги)'"
    if col_quantity == -1:
        return False, "Нет столбца 'Кол-во'"
    if col_summ == -1:
        return False, "Нет столбца 'Сумма'"

    # Forming lists , then bulk_create
    season_list = []  # list of instances
    capsule_list = []
    sku_list = []
    size_list = []

    seasons_checklist = []  # list for checking if instance already exists
    capsules_checklist = []
    skus_checklist = []

    row = row_header + 1  # start just below the header

    while row < sheet.max_row + 1:

        work_cell = sheet.cell(row=row, column=col_sku).value
        if not work_cell:
            row += 1
            continue

        work_cell = work_cell.strip()

        # Simple validation: if first 3 symbols are digits - this is sku
        if not str(work_cell)[:3].isdigit():
            row += 1
            continue

        # Check if everything is ok with values
        cell = sheet.cell(row=row, column=col_name).value  # name
        if cell:
            pos = cell.split(' ')
            skuname = pos[0]

        cell = sheet.cell(row=row, column=col_quantity).value  # quantity
        if cell:
            cell = str(cell)
            q_onway = string_to_integer(cell)
        else:
            q_onway = 0

        cell = sheet.cell(row=row, column=col_summ).value  # summ
        if cell:
            cell = str(cell)
            summ_onway = string_to_decimal(cell)
        else:
            summ_onway = 0

        if q_onway <= 0 or summ_onway <= 0:  # If quantity or summ <=0 - skip row
            row += 1
            continue

        # Parse the sku
        t = work_cell.find('-')
        if t != -1:
            sku_nosize = work_cell[:t]
            sizelong = work_cell[t + 1:]
        else:
            sku_nosize = work_cell
            sizelong = 'No size'

        season = work_cell[:5]
        if season[3:5] == 'GS' or season[3:5] == 'gs':
            pass  # SCHOOL, we take first 5 symbols
        else:
            season = season[0:3]  # NOT SCHOOL, we take first 3 symbols

        capsule = sku_nosize[:6]

        string_id_season = ''

        # Add Season
        # But first let's check if it already exists in database:
        if Season.objects.filter(season_firstletters=season, user=user_id).exists():
            one_entry = Season.objects.get(season_firstletters=season, user=user_id)
            string_id_season = one_entry.id

        else:
            if season not in seasons_checklist:
                seasons_checklist.append(season)
                try:
                    seasonname_from_config = seasons_configlist[season]
                except:
                    seasonname_from_config = '----'

                filename = 'images/' + season + '/' + season + '.jpg'
                image_season = STATIC_URL + filename

                if not os.path.isfile(STATIC_ROOT + '/' + filename):
                    image_season = ''

                # I have to assign id by myself, because whole database firstly create in memory, and after all - bulk_create.
                # And I need to organize ForeignKey relations, so I need id for that.
                # So, my id is CharField and looks like "onway-rownumber-user_id", i.e. onway-211-18.
                # Add "onway" because rownumber can be the same in existing database, so id-s can interfere.

                string_id_season = 'onway-' + str(row) + '-' + str(user_id.id)

                season_list.append(Season(season_firstletters=season,
                                          name=seasonname_from_config,
                                          img=image_season,
                                          user=user_id,
                                          id=string_id_season))

                # I create Season here, not later with bulk_create, because it needs to be existed for ForeignKey relations
                Season.objects.create(season_firstletters=season,
                                      name=seasonname_from_config,
                                      img=image_season,
                                      user=user_id,
                                      id=string_id_season)

        # Add Capsule
        # But first let's check if it already exists in database:
        if Capsule.objects.filter(capsule_firstletters=capsule, user=user_id).exists():
            one_entry = Capsule.objects.get(capsule_firstletters=capsule, user=user_id)
            string_id_capsule = one_entry.id
        else:
            if capsule not in capsules_checklist:

                capsules_checklist.append(capsule)
                try:
                    capsulename_from_config = capsules_configlist[capsule]
                except:
                    capsulename_from_config = '----'

                filename = 'images/' + season + '/' + capsule + '.jpg'
                image_capsule = STATIC_URL + filename

                if not os.path.isfile(STATIC_ROOT + '/' + filename):
                    image_capsule = ''

                season_id = string_id_season

                string_id_capsule = 'onway-' + str(row) + '-' + str(user_id.id)

                capsule_list.append(Capsule(capsule_firstletters=capsule,
                                            id=string_id_capsule,
                                            name=capsulename_from_config,
                                            img=image_capsule,
                                            user=user_id,
                                            season=Season(
                                                id=season_id)))

        # Add Sku
        # But first let's check if it already exists in database:
        if SKU.objects.filter(sku_firstletters=sku_nosize, user=user_id).exists():
            one_entry = SKU.objects.get(sku_firstletters=sku_nosize, user=user_id)
            string_id_sku = one_entry.id
        else:
            if sku_nosize not in skus_checklist:
                skus_checklist.append(sku_nosize)

                filename = 'images/' + season + '/' + capsule + '/' + sku_nosize + '.jpg'
                image_sku = STATIC_URL + filename

                if not os.path.isfile(STATIC_ROOT + '/' + filename):
                    image_sku = ''

                capsule_id = string_id_capsule

                string_id_sku = 'onway-' + str(row) + '-' + str(user_id.id)

                sku_list.append(SKU(name=skuname,
                                    sku_firstletters=sku_nosize,
                                    id=string_id_sku,
                                    capsule=Capsule(id=capsule_id),
                                    user=user_id,
                                    img=image_sku
                                    ))

        # Add Size
        if Size.objects.filter(size_long=sizelong, sku=string_id_sku, user=user_id).exists():
            # UPDATE this Size with quantity ONWAY and Summ ONWAY
            one_entry = Size.objects.get(size_long=sizelong, sku=string_id_sku, user=user_id)
            one_entry.quantity_onway = one_entry.quantity_onway + q_onway
            one_entry.costsumm_onway = one_entry.costsumm_onway + summ_onway
            one_entry.save()
        else:
            sku_id = string_id_sku

            s_s = get_size_short(sizelong)

            size_list.append(Size(size_long=sizelong,
                                  sku_full=work_cell,
                                  sku=SKU(id=sku_id),
                                  user=user_id,
                                  quantity_sold=0,
                                  sellsumm_sold=0,
                                  costsumm_sold=0,
                                  income=0,
                                  quantity_instock=0,
                                  costsumm_instock=0,
                                  size_short=s_s,
                                  quantity_onway=q_onway,
                                  costsumm_onway=summ_onway
                                  ))

        row += 1

    # By this moment we have lists: season_list, capsule_list, sku_list, size_list
    # In these lists those instances, that NOT exists in database.
    # If Size was already exist - we just added qquantity_onway and costsumm_onway

    #   ------------------ Now lets write to database new added instances ----------------
    if capsule_list:
        Capsule.objects.bulk_create(capsule_list)
    if sku_list:
        SKU.objects.bulk_create(sku_list)
    if size_list:
        Size.objects.bulk_create(size_list)

    return True, error_message
